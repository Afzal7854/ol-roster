from flask import Flask, request, send_file, render_template_string, session, redirect, url_for
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from functools import wraps
import re, io, os, json, hashlib, secrets
import psycopg2
from psycopg2.extras import RealDictCursor

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))

DATABASE_URL = os.environ.get('DATABASE_URL', '')

# ── DATABASE ─────────────────────────────────────────────────────
def get_db():
    return psycopg2.connect(DATABASE_URL)

def init_db():
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('''
            CREATE TABLE IF NOT EXISTS users (
                username VARCHAR(100) PRIMARY KEY,
                password VARCHAR(200) NOT NULL,
                role VARCHAR(20) DEFAULT 'user',
                active BOOLEAN DEFAULT TRUE,
                attempts INTEGER DEFAULT 0,
                locked BOOLEAN DEFAULT FALSE
            )
        ''')
        # Create default admin if not exists
        cur.execute("SELECT username FROM users WHERE username = 'admin'")
        if not cur.fetchone():
            cur.execute(
                "INSERT INTO users (username, password, role) VALUES (%s, %s, %s)",
                ('admin', hash_pw('OL@Admin2026'), 'admin')
            )
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"DB init error: {e}")

def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

def load_data():
    try:
        conn = get_db()
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute("SELECT * FROM users")
        rows = cur.fetchall()
        cur.close()
        conn.close()
        data = {}
        for row in rows:
            data[row['username']] = {
                'password': row['password'],
                'role': row['role'],
                'active': row['active'],
                'attempts': row['attempts'],
                'locked': row['locked']
            }
        return data
    except Exception as e:
        print(f"load_data error: {e}")
        return {'admin': {'password': hash_pw('OL@Admin2026'), 'role': 'admin', 'active': True, 'attempts': 0, 'locked': False}}

def save_user(username, data):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('''
            INSERT INTO users (username, password, role, active, attempts, locked)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (username) DO UPDATE SET
                password = EXCLUDED.password,
                role = EXCLUDED.role,
                active = EXCLUDED.active,
                attempts = EXCLUDED.attempts,
                locked = EXCLUDED.locked
        ''', (username, data['password'], data.get('role','user'),
              data.get('active',True), data.get('attempts',0), data.get('locked',False)))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"save_user error: {e}")

def delete_user(username):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("DELETE FROM users WHERE username = %s", (username,))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"delete_user error: {e}")

def update_user_field(username, field, value):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(f"UPDATE users SET {field} = %s WHERE username = %s", (value, username))
        conn.commit()
        cur.close()
        conn.close()
    except Exception as e:
        print(f"update_user error: {e}")

# ── RULES ────────────────────────────────────────────────────────
STATION_GF = {
    1:'GF-00003',2:'GF-00004',3:'GF-00005',4:'GF-00006',5:'GF-00007',
    6:'GF-00008',7:'GF-00009',8:'GF-00010',9:'GF-00011',10:'GF-00012',
    11:'GF-00013',12:'GF-00014',13:'GF-00015',14:'GF-00016',15:'GF-00017',
    16:'GF-00018',17:'GF-00019',18:'GF-00027',19:'GF-00028',20:'GF-00020',
    21:'GF-00021',22:'GF-00022',23:'GF-00023',24:'GF-00024',25:'GF-00025',26:'GF-00026'
}
SM_ODD = {
    1:1,2:3,3:3,4:5,5:5,6:7,7:7,8:9,9:9,10:11,11:11,12:13,13:13,
    14:15,15:15,16:17,17:17,18:19,19:19,20:21,21:21,22:23,23:23,24:25,25:25,26:26
}
SPECIAL_TS = {
    'SL':'TS-00068','CL':'TS-00069','AL':'TS-00070',
    'MI':'TS-00040','M':'TS-00061','D':'TS-00002',
    'TRG-DEPOT':'TS-00059','D-DEPOT':'TS-00002'
}
REST_CODES = {'REST','CPL','GHD'}

# ── DECORATORS ───────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session or session.get('role') != 'admin':
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated

# ── HTML ─────────────────────────────────────────────────────────
BASE_STYLE = '''
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
*{box-sizing:border-box;margin:0;padding:0}
:root{--bg:#0f1117;--sur:#1a1d27;--bdr:#2a2d3e;--acc:#4f8ef7;--acc2:#a78bfa;
      --ok:#34d399;--red:#f87171;--warn:#fbbf24;--txt:#e2e8f0;--mut:#64748b}
body{background:var(--bg);color:var(--txt);font-family:'Inter',sans-serif;min-height:100vh}
.container{max-width:600px;margin:0 auto;padding:40px 20px}
.center{display:flex;flex-direction:column;align-items:center}
.badge{display:inline-block;background:linear-gradient(135deg,#4f8ef720,#a78bfa20);
       border:1px solid #4f8ef740;color:var(--acc);font-size:11px;font-weight:600;
       letter-spacing:2px;text-transform:uppercase;padding:4px 14px;border-radius:20px;margin-bottom:14px}
h1{font-size:26px;font-weight:700;background:linear-gradient(135deg,#e2e8f0,#a78bfa);
   -webkit-background-clip:text;-webkit-text-fill-color:transparent;background-clip:text;margin-bottom:6px}
h2{font-size:18px;font-weight:600;color:var(--txt);margin-bottom:16px}
.sub{color:var(--mut);font-size:13px;margin-bottom:32px}
.card{background:var(--sur);border:1px solid var(--bdr);border-radius:14px;padding:24px;width:100%;margin-bottom:18px}
.ct{font-size:11px;font-weight:600;color:var(--mut);text-transform:uppercase;
    letter-spacing:1.5px;margin-bottom:16px;display:flex;align-items:center;gap:8px}
.ct::before{content:'';display:block;width:3px;height:13px;background:var(--acc);border-radius:2px}
.sn{display:inline-flex;align-items:center;justify-content:center;width:18px;height:18px;
    background:var(--acc);color:#fff;border-radius:50%;font-size:10px;font-weight:700}
.fl{display:flex;flex-direction:column;gap:6px;margin-bottom:14px}
.fl label{font-size:12px;color:var(--mut);font-weight:500;text-transform:uppercase;letter-spacing:.5px}
input[type=text],input[type=password],input[type=date]{background:var(--bg);border:1px solid var(--bdr);
  border-radius:8px;color:var(--txt);padding:10px 14px;font-size:13px;width:100%;outline:none;
  transition:border-color .2s;font-family:'Inter',sans-serif}
input:focus{border-color:var(--acc)}
.btn{width:100%;padding:13px;background:linear-gradient(135deg,var(--acc),var(--acc2));border:none;
     border-radius:10px;color:#fff;font-size:14px;font-weight:700;font-family:'Inter',sans-serif;
     cursor:pointer;transition:opacity .2s,transform .1s;margin-top:8px}
.btn:hover{opacity:.9;transform:translateY(-1px)}
.btn-sm{padding:7px 14px;border-radius:7px;border:none;font-size:12px;font-weight:600;
        font-family:'Inter',sans-serif;cursor:pointer;transition:all .2s}
.btn-ok{background:var(--ok);color:#000}
.btn-red{background:var(--red);color:#fff}
.btn-acc{background:var(--acc);color:#fff}
.btn-warn{background:var(--warn);color:#000}
.alert{padding:10px 14px;border-radius:8px;font-size:13px;margin-bottom:14px}
.alert-err{background:#f8717115;border:1px solid #f8717140;color:var(--red)}
.alert-ok{background:#34d39915;border:1px solid #34d39940;color:var(--ok)}
.alert-warn{background:#fbbf2415;border:1px solid #fbbf2440;color:var(--warn)}
.dz{border:2px dashed var(--bdr);border-radius:10px;padding:28px 20px;text-align:center;
    cursor:pointer;transition:all .2s;position:relative}
.dz:hover{border-color:var(--acc);background:#4f8ef708}
.dz input{position:absolute;inset:0;opacity:0;cursor:pointer;width:100%;height:100%}
.dz p{color:var(--mut);font-size:13px}
.fok{display:none;align-items:center;gap:8px;background:#34d39912;border:1px solid #34d39930;
     border-radius:7px;padding:8px 12px;margin-top:10px;color:var(--ok);font-size:12px}
.g2{display:grid;grid-template-columns:1fr 1fr;gap:12px}
.hint{grid-column:1/-1;font-size:11px;color:var(--mut);background:#4f8ef708;
      border:1px solid var(--bdr);border-radius:7px;padding:7px 11px}
table{width:100%;border-collapse:collapse;font-size:13px}
th{background:var(--bg);color:var(--mut);font-weight:600;padding:8px 12px;
   text-align:left;font-size:11px;text-transform:uppercase;letter-spacing:.5px}
td{padding:10px 12px;border-bottom:1px solid var(--bdr);color:var(--txt)}
tr:last-child td{border-bottom:none}
.badge-ok{background:#34d39920;color:var(--ok);padding:2px 8px;border-radius:12px;font-size:11px;font-weight:600}
.badge-red{background:#f8717120;color:var(--red);padding:2px 8px;border-radius:12px;font-size:11px;font-weight:600}
nav{background:var(--sur);border-bottom:1px solid var(--bdr);padding:14px 24px;
    display:flex;align-items:center;justify-content:space-between}
.nav-brand{font-weight:700;font-size:15px;color:var(--txt)}
.nav-brand span{color:var(--acc)}
.nav-right{display:flex;gap:10px;align-items:center}
.nav-user{font-size:12px;color:var(--mut)}
footer{text-align:center;color:var(--mut);font-size:11px;padding:24px;margin-top:8px}
.modal-bg{display:none;position:fixed;inset:0;background:#00000080;z-index:100;
          align-items:center;justify-content:center}
.modal-bg.show{display:flex}
.modal{background:var(--sur);border:1px solid var(--bdr);border-radius:14px;padding:28px;
       width:90%;max-width:420px}
.modal h3{font-size:16px;font-weight:600;margin-bottom:16px}
.modal-btns{display:flex;gap:10px;margin-top:16px;justify-content:flex-end}
</style>
'''

LOGIN_HTML = BASE_STYLE + '''
<div style="min-height:100vh;display:flex;align-items:center;justify-content:center;padding:20px">
<div style="width:100%;max-width:420px">
  <div style="text-align:center;margin-bottom:32px">
    <div class="badge">OL Operation Roster</div>
    <h1>Geo Fence Generator</h1>
    <p class="sub">Orange Line Metro Train — Lahore</p>
  </div>
  {% if error %}<div class="alert alert-err">{{ error }}</div>{% endif %}
  {% if locked %}<div class="alert alert-err">⛔ Account locked. Contact admin.</div>{% endif %}
  <div class="card">
    <div class="ct">Sign In</div>
    <form method="POST">
      <div class="fl"><label>Username</label>
        <input type="text" name="username" placeholder="Enter username" required autocomplete="off"></div>
      <div class="fl"><label>Password</label>
        <input type="password" name="password" placeholder="Enter password" required></div>
      {% if attempts and attempts > 0 %}
      <div class="alert alert-warn">⚠️ {{ attempts }} failed attempt(s). {{ 5 - attempts }} remaining.</div>
      {% endif %}
      <button type="submit" class="btn">Sign In</button>
    </form>
  </div>
  <footer>Orange Line Metro Train · Lahore · CPEC Project</footer>
</div></div>
'''

MAIN_HTML = BASE_STYLE + '''
<nav>
  <div class="nav-brand"><span>OL</span> Operation Roster</div>
  <div class="nav-right">
    <span class="nav-user">👤 {{ username }}</span>
    <a href="/logout"><button class="btn-sm btn-red">Logout</button></a>
    {% if role == 'admin' %}<a href="/admin"><button class="btn-sm btn-acc">⚙️ Admin</button></a>{% endif %}
  </div>
</nav>
<div class="container center" style="padding-top:28px">
  <div style="text-align:center;margin-bottom:28px">
    <div class="badge">Geo Fence Generator</div>
    <h1>OL Operation Roster</h1>
    <p class="sub">Generate attendance geo fence file from roster</p>
  </div>
  {% if error %}<div class="alert alert-err" style="width:100%">❌ {{ error }}</div>{% endif %}
  {% if missing_dates %}<div class="alert alert-warn" style="width:100%">
    ⚠️ <strong>Dates not found in roster:</strong> {{ missing_dates }}</div>{% endif %}
  <form method="POST" enctype="multipart/form-data" style="width:100%">
    <div class="card">
      <div class="ct"><span class="sn">1</span> Upload Roster File</div>
      <div class="dz" onclick="document.getElementById('rf').click()">
        <input type="file" name="roster" id="rf" accept=".xlsx,.xls" onchange="fileSelected(this)">
        <div style="font-size:30px;margin-bottom:8px">📂</div>
        <p><strong>Click to browse or drag &amp; drop</strong></p>
        <p style="margin-top:4px">OL Roster Excel file (.xlsx)</p>
      </div>
      <div class="fok" id="fok">✅ <span id="fn"></span></div>
    </div>
    <div class="card">
      <div class="ct"><span class="sn">2</span> Select Date Range</div>
      <div class="g2">
        <div class="fl" style="margin-bottom:0"><label>Start Date</label>
          <input type="date" name="start_date" id="sd" required></div>
        <div class="fl" style="margin-bottom:0"><label>End Date</label>
          <input type="date" name="end_date" id="ed" required></div>
        <div class="hint">ℹ️ Select any date range. Missing dates will show as warning.</div>
      </div>
    </div>
    <button type="submit" class="btn">🚀 Generate Geo Fence File</button>
  </form>
</div>
<footer>OL Operation Roster · Orange Line Metro · All data processed locally</footer>
<script>
function fileSelected(i){if(i.files&&i.files[0]){document.getElementById('fok').style.display='flex';document.getElementById('fn').textContent=i.files[0].name;}}
const t=new Date(),s=new Date(t);s.setDate(t.getDate()-6);
const f=d=>d.toISOString().split('T')[0];
document.getElementById('sd').value=f(s);document.getElementById('ed').value=f(t);
</script>
'''

ADMIN_HTML = BASE_STYLE + '''
<nav>
  <div class="nav-brand"><span>OL</span> Operation Roster</div>
  <div class="nav-right">
    <span class="nav-user">👤 {{ username }} (Admin)</span>
    <a href="/app"><button class="btn-sm btn-acc">🏠 App</button></a>
    <a href="/logout"><button class="btn-sm btn-red">Logout</button></a>
  </div>
</nav>
<div class="container" style="padding-top:28px">
  <h2>⚙️ Admin Panel — User Management</h2>
  {% if msg %}<div class="alert alert-ok">✅ {{ msg }}</div>{% endif %}
  {% if err %}<div class="alert alert-err">❌ {{ err }}</div>{% endif %}
  <div class="card">
    <div class="ct">Add New User</div>
    <form method="POST" action="/admin/add">
      <div class="g2">
        <div class="fl" style="margin-bottom:0"><label>Username</label>
          <input type="text" name="username" placeholder="e.g. sec_ali" required autocomplete="off"></div>
        <div class="fl" style="margin-bottom:0"><label>Password</label>
          <input type="password" name="password" placeholder="Set password" required></div>
      </div>
      <button type="submit" class="btn-sm btn-ok" style="margin-top:12px;padding:9px 20px">+ Add User</button>
    </form>
  </div>
  <div class="card">
    <div class="ct">All Users</div>
    <table>
      <tr><th>Username</th><th>Role</th><th>Status</th><th>Attempts</th><th>Actions</th></tr>
      {% for uname, udata in users.items() %}
      <tr>
        <td>{{ uname }}</td><td>{{ udata.role }}</td>
        <td>{% if udata.locked %}<span class="badge-red">Locked</span>
            {% else %}<span class="badge-ok">Active</span>{% endif %}</td>
        <td>{{ udata.attempts }}/5</td>
        <td style="display:flex;gap:6px;flex-wrap:wrap">
          {% if udata.locked %}
          <form method="POST" action="/admin/unlock"><input type="hidden" name="username" value="{{ uname }}">
          <button type="submit" class="btn-sm btn-ok">Unlock</button></form>{% endif %}
          <button onclick="showPw('{{ uname }}')" class="btn-sm btn-warn">Change PW</button>
          {% if uname != 'admin' %}
          <form method="POST" action="/admin/delete" onsubmit="return confirm('Delete?')">
          <input type="hidden" name="username" value="{{ uname }}">
          <button type="submit" class="btn-sm btn-red">Delete</button></form>{% endif %}
        </td>
      </tr>{% endfor %}
    </table>
  </div>
</div>
<div class="modal-bg" id="pwModal">
  <div class="modal"><h3>Change Password</h3>
    <form method="POST" action="/admin/change-pw">
      <input type="hidden" name="username" id="pwUser">
      <div class="fl"><label>New Password</label>
        <input type="password" name="new_password" placeholder="Enter new password" required></div>
      <div class="modal-btns">
        <button type="button" class="btn-sm btn-red" onclick="closePw()">Cancel</button>
        <button type="submit" class="btn-sm btn-ok">Save</button>
      </div>
    </form>
  </div>
</div>
<footer>OL Operation Roster · Admin Panel</footer>
<script>
function showPw(u){document.getElementById('pwUser').value=u;document.getElementById('pwModal').classList.add('show')}
function closePw(){document.getElementById('pwModal').classList.remove('show')}
</script>
'''

# ── ROUTES ───────────────────────────────────────────────────────
@app.route('/', methods=['GET'])
def index():
    if 'user' in session:
        return redirect(url_for('main'))
    return render_template_string(LOGIN_HTML)

@app.route('/', methods=['POST'])
def do_login():
    username = request.form.get('username','').strip().lower()
    password = request.form.get('password','')
    data = load_data()
    if username not in data:
        return render_template_string(LOGIN_HTML, error='Invalid username or password.')
    user = data[username]
    if user.get('locked'):
        return render_template_string(LOGIN_HTML, locked=True)
    if user['password'] == hash_pw(password):
        update_user_field(username, 'attempts', 0)
        session.permanent = True
        session['user'] = username
        session['role'] = user.get('role','user')
        return redirect(url_for('main'))
    else:
        new_attempts = user.get('attempts',0) + 1
        update_user_field(username, 'attempts', new_attempts)
        if new_attempts >= 5:
            update_user_field(username, 'locked', True)
            return render_template_string(LOGIN_HTML, locked=True)
        return render_template_string(LOGIN_HTML, error='Invalid username or password.', attempts=new_attempts)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/app', methods=['GET','POST'])
@login_required
def main():
    if request.method == 'GET':
        return render_template_string(MAIN_HTML, username=session['user'], role=session.get('role','user'))
    try:
        if 'roster' not in request.files or request.files['roster'].filename == '':
            return render_template_string(MAIN_HTML, username=session['user'], role=session.get('role','user'), error='Please upload a roster file.')
        start_str = request.form.get('start_date')
        end_str   = request.form.get('end_date')
        if not start_str or not end_str:
            return render_template_string(MAIN_HTML, username=session['user'], role=session.get('role','user'), error='Please select dates.')
        start = datetime.strptime(start_str, '%Y-%m-%d')
        end   = datetime.strptime(end_str,   '%Y-%m-%d')
        if start > end:
            return render_template_string(MAIN_HTML, username=session['user'], role=session.get('role','user'), error='Start date must be before end date.')
        roster_bytes = request.files['roster'].read()
        buf, count, missing = process_roster(roster_bytes, start, end)
        if count == 0:
            return render_template_string(MAIN_HTML, username=session['user'], role=session.get('role','user'), error='No data found. Check roster file and dates.')
        fname = f'Geo_Fence_{start_str}_to_{end_str}.xlsx'
        return send_file(buf, as_attachment=True, download_name=fname,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return render_template_string(MAIN_HTML, username=session['user'], role=session.get('role','user'), error=f'Error: {str(e)}')

@app.route('/admin')
@admin_required
def admin():
    return render_template_string(ADMIN_HTML, username=session['user'], users=load_data())

@app.route('/admin/add', methods=['POST'])
@admin_required
def admin_add():
    uname = request.form.get('username','').strip().lower()
    pw    = request.form.get('password','')
    data  = load_data()
    if not uname or not pw:
        return render_template_string(ADMIN_HTML, username=session['user'], users=data, err='Username and password required.')
    if uname in data:
        return render_template_string(ADMIN_HTML, username=session['user'], users=data, err=f'User "{uname}" already exists.')
    save_user(uname, {'password':hash_pw(pw),'role':'user','active':True,'attempts':0,'locked':False})
    return render_template_string(ADMIN_HTML, username=session['user'], users=load_data(), msg=f'User "{uname}" added.')

@app.route('/admin/delete', methods=['POST'])
@admin_required
def admin_delete():
    uname = request.form.get('username')
    if uname == 'admin':
        return render_template_string(ADMIN_HTML, username=session['user'], users=load_data(), err='Cannot delete admin.')
    delete_user(uname)
    return render_template_string(ADMIN_HTML, username=session['user'], users=load_data(), msg=f'User "{uname}" deleted.')

@app.route('/admin/unlock', methods=['POST'])
@admin_required
def admin_unlock():
    uname = request.form.get('username')
    update_user_field(uname, 'locked', False)
    update_user_field(uname, 'attempts', 0)
    return render_template_string(ADMIN_HTML, username=session['user'], users=load_data(), msg=f'User "{uname}" unlocked.')

@app.route('/admin/change-pw', methods=['POST'])
@admin_required
def admin_change_pw():
    uname  = request.form.get('username')
    new_pw = request.form.get('new_password','')
    if new_pw:
        save_user(uname, {'password':hash_pw(new_pw),'role':load_data().get(uname,{}).get('role','user'),'active':True,'attempts':0,'locked':False})
    return render_template_string(ADMIN_HTML, username=session['user'], users=load_data(), msg=f'Password changed for "{uname}".')

# ── GEO FENCE LOGIC ──────────────────────────────────────────────
def stn_num(s):
    if not s: return None
    m = re.search(r'(\d+)', str(s))
    return int(m.group(1)) if m else None

def get_ts(duty):
    if not duty: return None
    s = str(duty).strip().upper()
    if not s or s in REST_CODES: return None
    base = s.split('-')[0]
    if base == 'TRG' and s != 'TRG-DEPOT': return 'TS-00002'
    if s in SPECIAL_TS: return SPECIAL_TS[s]
    last = base[-1]
    if last == '1': return 'TS-00003'
    if last == '2': return 'TS-00004'
    return None

def get_nums(s):
    return [int(p) for p in s.upper().split('-')[1:] if re.match(r'^\d+$',p)]

def sm_gf(nums, own_stn):
    if not nums: return STATION_GF.get(SM_ODD.get(own_stn,own_stn),'')
    odds=[n for n in nums if n%2!=0]
    chosen=odds[0] if odds else nums[-1]
    return STATION_GF.get(SM_ODD.get(chosen,chosen),'')

def sec_gf(nums, own_stn):
    if not nums: return STATION_GF.get(own_stn,'')
    return STATION_GF.get(nums[-1],'')

def parse_sm(duty, own_stn):
    if not duty: return None,''
    s=str(duty).strip();su=s.upper()
    if su in REST_CODES: return None,''
    ts=get_ts(s)
    if su=='TRG-DEPOT': return 'TS-00059','GF-00001'
    if su=='D-DEPOT': return 'TS-00002','GF-00001'
    m=re.match(r'^TRG-(\d+)$',su)
    if m: return 'TS-00002',STATION_GF.get(int(m.group(1)),'')
    nums=get_nums(su)
    return ts,(sm_gf(nums,own_stn) if ts else '')

def parse_sec(duty, own_stn):
    if not duty: return None,''
    s=str(duty).strip();su=s.upper()
    if su in REST_CODES: return None,''
    ts=get_ts(s)
    if su=='TRG-DEPOT': return 'TS-00059','GF-00001'
    if su=='D-DEPOT': return 'TS-00002','GF-00001'
    m=re.match(r'^TRG-(\d+)$',su)
    if m: return 'TS-00002',STATION_GF.get(int(m.group(1)),'')
    nums=get_nums(su)
    return ts,(sec_gf(nums,own_stn) if ts else '')

def parse_sec2sm(duty, col_d):
    if not duty: return None,''
    s=str(duty).strip();su=s.upper()
    if su in REST_CODES: return None,''
    ts=get_ts(s)
    if su=='TRG-DEPOT': return 'TS-00059','GF-00001'
    if su=='D-DEPOT': return 'TS-00002','GF-00001'
    m=re.match(r'^TRG-(\d+)$',su)
    if m: return 'TS-00002',STATION_GF.get(int(m.group(1)),'')
    duty_nums=get_nums(su)
    base=su.split('-')[0]
    if duty_nums:
        gf=(sm_gf(duty_nums,stn_num(col_d)) if (base.startswith('A') or base.startswith('KA')) else sec_gf(duty_nums,stn_num(col_d)))
    else:
        if col_d:
            cdn=[int(p.strip()) for p in str(col_d).split('-') if re.match(r'^\d+$',p.strip())]
            gf=sm_gf(cdn,cdn[0] if cdn else stn_num(col_d))
        else: gf=''
    return ts,(gf if ts else '')

def parse_mgt(duty, location):
    if not duty: return None,''
    s=str(duty).strip();su=s.upper()
    if su in REST_CODES: return None,''
    ts=get_ts(s)
    loc=str(location or '').strip().upper()
    if su=='D-DEPOT': return 'TS-00002','GF-00001'
    nums=get_nums(su)
    if loc=='OCC GF' and not nums: return ts or 'TS-00002','GF-00001'
    if nums:
        base=su.split('-')[0]
        gf=(sm_gf(nums,stn_num(location)) if (base.startswith('A') or base.startswith('KA')) else sec_gf(nums,stn_num(location)))
        return ts,(gf if ts else '')
    sn=stn_num(location)
    return ts,(STATION_GF.get(sn,'') if ts and sn else '')

def process_roster(roster_bytes, start_date, end_date):
    target_dates=[]
    d=start_date
    while d<=end_date:
        target_dates.append(d)
        d+=timedelta(days=1)
    wb=load_workbook(io.BytesIO(roster_bytes),read_only=True)
    olt_map={}
    found_dates=set()
    SHEET_CFG={
        'SM':{'oltCol':1,'stnCol':0,'type':'SM'},
        'SEC':{'oltCol':1,'stnCol':0,'type':'SEC'},
        'SA':{'oltCol':1,'stnCol':0,'type':'SA'},
        'SEC to SM':{'oltCol':1,'stnCol':0,'colD':3,'type':'SEC2SM'},
        'Management':{'oltCol':3,'stnCol':1,'type':'MGT'},
    }
    for sheet,cfg in SHEET_CFG.items():
        if sheet not in wb.sheetnames: continue
        ws=wb[sheet]
        rows=list(ws.iter_rows(values_only=True))
        date_cols={}
        for i,v in enumerate(rows[2]):
            if isinstance(v,datetime) and v in target_dates:
                date_cols[v.strftime('%Y-%m-%d')]=i
                found_dates.add(v.strftime('%Y-%m-%d'))
        last_stn=None
        for row in rows[3:]:
            if not row: continue
            sv=row[cfg['stnCol']]
            if sv is not None and sv!='': last_stn=sv
            olt=row[cfg['oltCol']]
            if not olt or not str(olt).startswith('OLT'): continue
            olt_str=str(olt).strip()
            if olt_str not in olt_map: olt_map[olt_str]={}
            # SEC2SM: find correct station col for this week's dates
            col_d = None
            if 'colD' in cfg:
                # Find station col = the 'Station' header col just before the date cols for this week
                for dc_str, dc_idx in date_cols.items():
                    # station col is the col with 'Station' header closest before dc_idx
                    best_stn_col = cfg['colD']  # default col 3
                    for hi, hv in enumerate(rows[1]):
                        if hv == 'Station' and hi < dc_idx:
                            best_stn_col = hi
                    col_d = row[best_stn_col] if best_stn_col < len(row) else None
                    break  # use first date's station col as representative
            # Build station col map for SEC2SM: each date -> its station col
            stn_col_map = {}
            if cfg['type'] == 'SEC2SM':
                stn_cols = [i for i,v in enumerate(rows[1]) if v == 'Station']
                date_col_list = sorted(date_cols.items(), key=lambda x: x[1])
                for d_str, dc_idx in date_col_list:
                    best = cfg['colD']
                    for sc in stn_cols:
                        if sc < dc_idx:
                            best = sc
                    stn_col_map[d_str] = best

            for d_str,ci in date_cols.items():
                if ci>=len(row): continue
                duty=row[ci]
                if not duty or str(duty).startswith('='): continue
                duty_s=str(duty).strip()
                if not duty_s: continue
                if d_str in olt_map[olt_str] and olt_map[olt_str][d_str].get('ts'): continue
                if cfg['type']=='SM': ts,gf=parse_sm(duty_s,stn_num(last_stn))
                elif cfg['type'] in ('SEC','SA'): ts,gf=parse_sec(duty_s,stn_num(last_stn))
                elif cfg['type']=='SEC2SM':
                    week_col_d = row[stn_col_map.get(d_str, cfg['colD'])] if stn_col_map else col_d
                    ts,gf=parse_sec2sm(duty_s,week_col_d)
                else: ts,gf=parse_mgt(duty_s,last_stn)
                olt_map[olt_str][d_str]={'ts':ts,'gf':gf}
    missing=[d.strftime('%Y-%m-%d') for d in target_dates if d.strftime('%Y-%m-%d') not in found_dates]
    employees=sorted(olt_map.items(),key=lambda x:int(x[0].replace('OLT-','')))
    HDR_FILL=PatternFill('solid',start_color='CC0000',end_color='CC0000')
    TS_FILL=PatternFill('solid',start_color='C00000',end_color='C00000')
    GF_FILL=PatternFill('solid',start_color='00B050',end_color='00B050')
    HDR_FONT=Font(name='Arial',size=9,bold=True,color='FFFFFF')
    SUB_FONT=Font(name='Arial',size=9,color='000000')
    DAT_FONT=Font(name='Arial',size=9,color='000000')
    CTR=Alignment(horizontal='center',vertical='center')
    LEFT=Alignment(horizontal='left',vertical='center')
    thin=Side(style='thin',color='CCCCCC')
    BRD=Border(top=thin,bottom=thin,left=thin,right=thin)
    wb_out=Workbook()
    ws_out=wb_out.active
    ws_out.title='Roster With Geo Fence'
    CPD=4
    ws_out.column_dimensions['A'].width=32.88
    ws_out.row_dimensions[1].height=27.75
    ws_out.row_dimensions[2].height=15.75
    for di in range(len(target_dates)):
        base=di*CPD
        ws_out.column_dimensions[get_column_letter(2+base)].width=9.44
        ws_out.column_dimensions[get_column_letter(3+base)].width=18.0
        ws_out.column_dimensions[get_column_letter(4+base)].width=9.44
        ws_out.column_dimensions[get_column_letter(5+base)].width=13.0
    c1=ws_out.cell(1,1,'Employee Code')
    c1.fill=HDR_FILL;c1.font=HDR_FONT;c1.alignment=CTR;c1.border=BRD
    ws_out.merge_cells(start_row=1,start_column=1,end_row=2,end_column=1)
    for di,dt in enumerate(target_dates):
        col=2+di*CPD
        dc=ws_out.cell(1,col,dt.strftime('%Y-%m-%d'))
        dc.fill=HDR_FILL;dc.font=HDR_FONT;dc.alignment=CTR;dc.border=BRD
        ws_out.merge_cells(start_row=1,start_column=col,end_row=1,end_column=col+CPD-1)
        for lbl2,fill2 in [('Time Slot',TS_FILL),('Geofence',GF_FILL),('Rest Day',GF_FILL),('Full Overtime',GF_FILL)]:
            sc=ws_out.cell(2,col,lbl2)
            sc.fill=fill2;sc.font=SUB_FONT;sc.alignment=CTR;sc.border=BRD
            col+=1
    for ri,(olt,dates) in enumerate(employees,start=3):
        ws_out.row_dimensions[ri].height=15.0
        ec=ws_out.cell(ri,1,olt)
        ec.font=DAT_FONT;ec.alignment=LEFT;ec.border=BRD
        for di,dt in enumerate(target_dates):
            col=2+di*CPD
            d_str=dt.strftime('%Y-%m-%d')
            entry=dates.get(d_str,{})
            ts=entry.get('ts') or ''
            gf=entry.get('gf') or ''
            working=bool(ts)
            for val in [ts,gf,'FALSE' if working else '','FALSE' if working else '']:
                c=ws_out.cell(ri,col,val)
                c.font=DAT_FONT;c.alignment=CTR;c.border=BRD
                col+=1
    buf=io.BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf,len(employees),missing

if __name__=='__main__':
    if DATABASE_URL:
        init_db()
    import webbrowser,threading
    def open_browser():
        import time;time.sleep(1.5)
        webbrowser.open('http://localhost:5000')
    threading.Thread(target=open_browser,daemon=True).start()
    print("="*55)
    print("  OL Operation Roster — Geo Fence Generator")
    print("  URL: http://localhost:5000")
    print("  Admin: http://localhost:5000/admin")
    print("="*55)
    app.run(debug=False,port=5000,host='0.0.0.0')